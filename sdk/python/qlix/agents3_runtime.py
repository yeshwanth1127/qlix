"""Hybrid local execution via gui-agents S3 (Agent-S3), not Luna tools."""

from __future__ import annotations

import asyncio
import json
import os
import platform
import re
import subprocess
import sys
from pathlib import Path
from typing import Any, Awaitable, Callable

from .agents3_proxy import Agents3RunContext, resolve_s3_engine_params
from .identity import AgentIdentity

Agents3Executor = Callable[[str], Awaitable[str]]

TOOL_SCOPE_MAP: dict[str, tuple[str, ...]] = {
    "s3_read_file": ("system.file_read",),
    "s3_write_file": ("system.file_write",),
    "s3_list_dir": ("system.file_read",),
    "s3_bash": ("system.file_write",),
    "s3_python": ("system.file_write",),
    "s3_code_task": ("system.file_write",),
    "s3_open_file": ("system.file_read",),
    "s3_create_pdf": ("system.file_write",),
    "s3_create_xlsx": ("system.file_write",),
    "s3_send_whatsapp_document": ("system.file_read",),
    "gui_control": ("system.gui_control",),
}

READ_ONLY_S3_TOOLS = frozenset({"s3_read_file", "s3_list_dir", "s3_open_file"})
WRITE_S3_TOOLS = frozenset(
    {
        "s3_write_file",
        "s3_bash",
        "s3_python",
        "s3_code_task",
        "s3_create_pdf",
        "s3_create_xlsx",
    }
)

# Binary office/document formats that must NOT be produced by s3_write_file (which
# only writes UTF-8 text — doing so yields a corrupt file the app can't open).
_BINARY_DOC_SUFFIXES = {".xlsx", ".xls", ".docx", ".pptx", ".pdf"}

_WRITE_INTENT = re.compile(
    r"\b(write|save|overwrite|append|export|"
    r"create\s+(?:a\s+)?(?:new\s+)?(?:file|pdf|document|doc|report|invoice)|"
    r"make\s+(?:a\s+)?(?:pdf|document|report)|"
    r"generate\s+.*\.(?:txt|log|csv|json|md|pdf|docx|xlsx))\b",
    re.IGNORECASE,
)
_READ_INTENT = re.compile(
    r"\b(read|open|view|show|list|inspect|summarize|summary|review|scan|tail)\b",
    re.IGNORECASE,
)

# A produced/delivered artifact means the task is NOT read-only even when the
# instruction also contains a read verb (e.g. "read the doc and make a pdf").
_OUTPUT_INTENT = re.compile(
    r"\b(pdf|docx|xlsx|spreadsheet|report|invoice|document|"
    r"send|share|deliver|attach|download|export|email|whatsapp)\b",
    re.IGNORECASE,
)


def is_read_only_file_intent(instruction: str) -> bool:
    """True when the user asked to read/open/review files but not create or modify them.

    A request that also produces or delivers an artifact (pdf/report/whatsapp/send/…)
    is never read-only, even if it contains a read verb like "summarize".
    """
    text = (instruction or "").strip()
    if not text:
        return False
    if _WRITE_INTENT.search(text) or _OUTPUT_INTENT.search(text):
        return False
    return _READ_INTENT.search(text) is not None

LOCAL_TOOL_IDS = (
    "s3_read_file",
    "s3_write_file",
    "s3_list_dir",
    "s3_open_file",
    "s3_bash",
    "s3_python",
    "s3_create_pdf",
    "s3_create_xlsx",
    "s3_send_whatsapp_document",
)
CODE_TOOL_IDS = ("s3_bash", "s3_python", "s3_code_task")
GUI_TOOL_IDS = ("gui_control",)


def _granted(identity: AgentIdentity) -> set[str]:
    return set(identity.permission_scopes) | set(identity.always_scopes)


def _filter_tools(
    tool_ids: tuple[str, ...],
    identity: AgentIdentity,
    skill_filter: list[str] | None,
    *,
    instruction: str | None = None,
) -> list[str]:
    granted = _granted(identity)
    out: list[str] = []
    filt = {str(s).strip() for s in (skill_filter or []) if str(s).strip()} if skill_filter else None
    read_only = is_read_only_file_intent(instruction or "")
    for tid in tool_ids:
        if read_only and tid in WRITE_S3_TOOLS:
            continue
        scopes = TOOL_SCOPE_MAP.get(tid, (tid,))
        if any(s not in granted for s in scopes):
            continue
        if filt and any("." in s for s in filt):
            if not any(s in filt for s in scopes):
                continue
        elif filt and tid not in filt:
            continue
        out.append(tid)
    return out


def diagnose_local_tools(
    identity: AgentIdentity,
    *,
    groups: tuple[str, ...],
    skill_filter: list[str] | None,
    instruction: str | None,
) -> dict[str, Any]:
    """Explain, per local tool, whether it is offered and (if not) why.

    Used by the runners to emit a debug event so a missing write/PDF tool is no
    longer silent. Mirrors the exact gating used by ``_filter_tools`` /
    ``openai_agents3_tool_definitions`` so the report can't drift from reality.
    """
    granted = _granted(identity)
    read_only = is_read_only_file_intent(instruction or "")
    filt = (
        {str(s).strip() for s in (skill_filter or []) if str(s).strip()}
        if skill_filter
        else None
    )
    scoped_filter = bool(filt and any("." in s for s in filt))

    # Which id pools are in play given the selected groups (matches the
    # openai_agents3_tool_definitions / build_agents3_executors logic).
    candidate_ids: list[str] = []
    if "files" in groups:
        candidate_ids.extend(LOCAL_TOOL_IDS)
    if "code" in groups and not read_only:
        candidate_ids.extend(t for t in CODE_TOOL_IDS if t not in candidate_ids)
    if "gui" in groups:
        candidate_ids.extend(t for t in GUI_TOOL_IDS if t not in candidate_ids)

    decisions: list[dict[str, Any]] = []
    offered: list[str] = []
    for tid in dict.fromkeys(LOCAL_TOOL_IDS + CODE_TOOL_IDS + GUI_TOOL_IDS):
        scopes = TOOL_SCOPE_MAP.get(tid, (tid,))
        missing = [s for s in scopes if s not in granted]
        reason = None
        if tid not in candidate_ids:
            reason = "group_not_selected_or_read_only"
        elif read_only and tid in WRITE_S3_TOOLS:
            reason = "blocked_by_read_only_intent"
        elif missing:
            reason = f"missing_scope:{','.join(missing)}"
        elif scoped_filter and not any(s in filt for s in scopes):
            reason = "skill_scope_filter_excluded"
        elif filt and not scoped_filter and tid not in filt:
            reason = "skill_tool_filter_excluded"
        if reason is None:
            offered.append(tid)
        decisions.append({"tool": tid, "offered": reason is None, "reason": reason})

    return {
        "granted_scopes": sorted(granted),
        "read_only_intent": read_only,
        "skill_filter": sorted(filt) if filt else None,
        "groups": list(groups),
        "instruction_preview": (instruction or "")[:200],
        "offered_tools": offered,
        "decisions": decisions,
    }


def _open_path_on_system(
    path: Path,
    *,
    mode: str = "default",
    application: str | None = None,
) -> tuple[bool, str]:
    """Launch file/folder on the user's desktop (native OS UI)."""
    if not path.exists():
        return False, f"Path not found: {path}"

    mode = (mode or "default").strip().lower()
    app = (application or "").strip()
    system = platform.system()

    try:
        if system == "Windows":
            p = str(path.resolve())
            if mode == "folder":
                target = p if path.is_dir() else str(path.parent.resolve())
                proc = subprocess.run(
                    ["explorer.exe", target],
                    capture_output=True,
                    text=True,
                    timeout=30,
                )
            elif mode == "reveal":
                proc = subprocess.run(
                    ["explorer.exe", "/select,", p],
                    capture_output=True,
                    text=True,
                    timeout=30,
                )
            elif app:
                proc = subprocess.run(
                    [
                        "powershell",
                        "-NoProfile",
                        "-Command",
                        f"Start-Process -FilePath {json.dumps(app)} -ArgumentList {json.dumps(p)}",
                    ],
                    capture_output=True,
                    text=True,
                    timeout=30,
                )
            else:
                proc = subprocess.run(
                    [
                        "powershell",
                        "-NoProfile",
                        "-Command",
                        f"Start-Process -FilePath {json.dumps(p)}",
                    ],
                    capture_output=True,
                    text=True,
                    timeout=30,
                )
            if proc.returncode != 0:
                err = (proc.stderr or proc.stdout or "").strip()
                return False, err or f"explorer/Start-Process exited {proc.returncode}"
            return True, f"Opened on screen: {p}"

        if system == "Darwin":
            cmd = ["open"]
            if mode == "reveal":
                cmd.append("-R")
            elif app:
                cmd.extend(["-a", app])
            cmd.append(str(path.resolve()))
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            if proc.returncode != 0:
                return False, (proc.stderr or proc.stdout or "").strip()
            return True, f"Opened on screen: {path}"

        # Linux and other Unix
        target = path.resolve()
        if mode == "folder" and path.is_file():
            target = path.parent
        proc = subprocess.run(
            ["xdg-open", str(target)],
            capture_output=True,
            text=True,
            timeout=30,
        )
        if proc.returncode != 0:
            return False, (proc.stderr or proc.stdout or "").strip()
        return True, f"Opened on screen: {target}"
    except subprocess.TimeoutExpired:
        return False, "Timed out launching application"
    except Exception as exc:
        return False, str(exc)


def _run_bash(code: str, *, timeout: int = 120) -> dict[str, Any]:
    """gui-agents LocalEnv bash, with Windows PowerShell support."""
    try:
        if platform.system() == "Windows":
            proc = subprocess.run(
                ["powershell", "-NoProfile", "-NonInteractive", "-Command", code],
                capture_output=True,
                text=True,
                timeout=timeout,
            )
        else:
            proc = subprocess.run(
                ["/bin/bash", "-lc", code],
                capture_output=True,
                text=True,
                timeout=timeout,
            )
        output = (proc.stdout or "") + (proc.stderr or "")
        return {
            "status": "ok" if proc.returncode == 0 else "error",
            "returncode": proc.returncode,
            "output": output,
        }
    except subprocess.TimeoutExpired as exc:
        return {"status": "error", "returncode": -1, "output": "", "error": str(exc)}
    except Exception as exc:
        return {"status": "error", "returncode": -1, "output": "", "error": str(exc)}


def _run_python(code: str) -> dict[str, Any]:
    try:
        from gui_agents.s3.utils.local_env import LocalEnv

        return LocalEnv().controller.run_python_script(code)
    except ImportError:
        proc = subprocess.run(
            [sys.executable, "-c", code],
            capture_output=True,
            text=True,
            timeout=120,
        )
        return {
            "status": "ok" if proc.returncode == 0 else "error",
            "returncode": proc.returncode,
            "output": (proc.stdout or "") + (proc.stderr or ""),
        }



def _slugify(text: str, *, fallback: str = "document") -> str:
    slug = re.sub(r"[^a-zA-Z0-9._-]+", "-", (text or "")).strip("-")
    return slug[:80] or fallback


def _create_pdf_file(
    *, title: str, content: str, output_path: str | None
) -> tuple[bool, str]:
    """Render text/markdown to a real PDF on the local filesystem.

    Returns ``(ok, message)`` where ``message`` is the absolute path on success
    or an error description on failure.
    """
    import tempfile

    if not (content.strip() or title.strip()):
        return False, "content is required to create a PDF"

    if output_path:
        out = Path(output_path).expanduser()
        if out.suffix.lower() != ".pdf":
            out = out.with_suffix(".pdf")
    else:
        base = Path.home() / "Documents"
        if not base.is_dir():
            base = Path(tempfile.gettempdir())
        out = base / f"{_slugify(title)}.pdf"

    try:
        out.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        return False, f"Cannot create output directory: {exc}"

    try:
        import html

        from reportlab.lib.pagesizes import LETTER
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            ListFlowable,
            ListItem,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
        )
    except ImportError:
        return (
            False,
            "reportlab is not installed. Install the hybrid extras: "
            "pip install 'qlix[hybrid]' (or pip install reportlab).",
        )

    styles = getSampleStyleSheet()

    def _md_inline(text: str) -> str:
        esc = html.escape(text)
        esc = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", esc)
        esc = re.sub(r"(?<![*\w])\*(?!\s)(.+?)(?<!\s)\*", r"<i>\1</i>", esc)
        return esc

    story: list[Any] = []
    if title.strip():
        story.append(Paragraph(_md_inline(title.strip()), styles["Title"]))
        story.append(Spacer(1, 0.2 * inch))

    blocks = re.split(r"\n\s*\n", content.replace("\r\n", "\n"))
    for block in blocks:
        block = block.strip("\n")
        if not block.strip():
            continue
        lines = block.split("\n")
        bullet_lines = [ln for ln in lines if ln.lstrip().startswith(("- ", "* "))]
        if bullet_lines and len(bullet_lines) == len(lines):
            items = [
                ListItem(Paragraph(_md_inline(ln.lstrip()[2:]), styles["BodyText"]))
                for ln in lines
            ]
            story.append(ListFlowable(items, bulletType="bullet"))
            story.append(Spacer(1, 6))
            continue
        if block.startswith("### "):
            story.append(Paragraph(_md_inline(block[4:]), styles["Heading3"]))
        elif block.startswith("## "):
            story.append(Paragraph(_md_inline(block[3:]), styles["Heading2"]))
        elif block.startswith("# "):
            story.append(Paragraph(_md_inline(block[2:]), styles["Heading1"]))
        else:
            text = "<br/>".join(_md_inline(ln) for ln in lines)
            story.append(Paragraph(text, styles["BodyText"]))
        story.append(Spacer(1, 6))

    try:
        doc = SimpleDocTemplate(
            str(out),
            pagesize=LETTER,
            leftMargin=0.9 * inch,
            rightMargin=0.9 * inch,
            topMargin=0.9 * inch,
            bottomMargin=0.9 * inch,
            title=title.strip() or out.stem,
        )
        doc.build(story)
    except Exception as exc:
        return False, f"PDF generation error: {exc}"

    return True, f"Created PDF: {out.resolve()}"


def _coerce_rows(data: Any) -> list[list[Any]] | None:
    """Normalize accepted xlsx inputs into a list of rows (list of cells)."""
    if isinstance(data, str):
        text = data.strip()
        if not text:
            return None
        # Try JSON first (list of lists or list of dicts), then CSV.
        try:
            parsed = json.loads(text)
        except (json.JSONDecodeError, ValueError):
            parsed = None
        if parsed is not None:
            return _coerce_rows(parsed)
        import csv
        import io

        return [row for row in csv.reader(io.StringIO(text))]
    if isinstance(data, list):
        if not data:
            return None
        if all(isinstance(r, dict) for r in data):
            headers: list[str] = []
            for r in data:
                for k in r.keys():
                    if k not in headers:
                        headers.append(str(k))
            rows: list[list[Any]] = [headers]
            for r in data:
                rows.append([r.get(h, "") for h in headers])
            return rows
        out: list[list[Any]] = []
        for r in data:
            if isinstance(r, list):
                out.append(list(r))
            else:
                out.append([r])
        return out
    return None


def _create_xlsx_file(
    *, title: str, rows: Any, sheet_name: str, output_path: str | None
) -> tuple[bool, str]:
    """Write a real .xlsx workbook using openpyxl.

    ``rows`` may be a list of lists, a list of dicts (keys become the header row),
    a JSON string of either, or CSV text. Returns ``(ok, message)``.
    """
    import tempfile

    table = _coerce_rows(rows)
    if not table:
        return False, "rows is required to create a spreadsheet (list of lists, list of dicts, JSON, or CSV)"

    if output_path:
        out = Path(output_path).expanduser()
        if out.suffix.lower() != ".xlsx":
            out = out.with_suffix(".xlsx")
    else:
        base = Path.home() / "Documents"
        if not base.is_dir():
            base = Path(tempfile.gettempdir())
        out = base / f"{_slugify(title or sheet_name or 'spreadsheet')}.xlsx"

    try:
        out.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        return False, f"Cannot create output directory: {exc}"

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return (
            False,
            "openpyxl is not installed. Install the hybrid extras: "
            "pip install 'qlix[hybrid]' (or pip install openpyxl).",
        )

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = (sheet_name or "Sheet1")[:31]
        for r_idx, row in enumerate(table, start=1):
            for c_idx, cell in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=cell)
        # Bold the header row and set a reasonable column width.
        if table:
            for c_idx in range(1, len(table[0]) + 1):
                ws.cell(row=1, column=c_idx).font = Font(bold=True)
            for c_idx in range(1, max(len(r) for r in table) + 1):
                width = max(
                    (len(str(r[c_idx - 1])) for r in table if c_idx - 1 < len(r)),
                    default=10,
                )
                ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = min(
                    max(width + 2, 10), 60
                )
        wb.save(str(out))
    except Exception as exc:
        return False, f"XLSX generation error: {exc}"

    return True, f"Created spreadsheet: {out.resolve()}"


async def _send_whatsapp_document(
    file_path: str,
    file_name: str | None,
    *,
    identity: AgentIdentity,
    agents3_context: Agents3RunContext | None,
) -> str:
    """Deliver a local file to the user's linked WhatsApp via the Qlix backend.

    The backend (which holds the WhatsApp service secret) forwards the request to
    the qlix-whatsapp-service ``/send-document`` endpoint.
    """
    if not file_path:
        return "[failed] file_path is required"
    path = Path(file_path).expanduser()
    if not path.is_file():
        return f"[failed] File not found: {path}"

    backend_url, agent_id, runner_token, _ = _s3_run_context(identity, agents3_context)
    if not (backend_url and agent_id and runner_token):
        return (
            "[failed] WhatsApp send is only available for hybrid runs with a runner token "
            "(missing backend_url/agent_id/runner_token)."
        )

    body: dict[str, Any] = {
        "file_path": str(path.resolve()),
        "file_name": file_name or path.name,
    }
    run_id = agents3_context.run_id if agents3_context else None
    if run_id:
        body["runId"] = run_id

    url = f"{backend_url.rstrip('/')}/api/v1/agents/{agent_id}/tools/whatsapp/send-document"
    headers = {"X-QLIX-Runner-Token": runner_token, "Content-Type": "application/json"}
    try:
        import httpx

        def _post() -> tuple[int, str]:
            with httpx.Client(timeout=60.0) as client:
                resp = client.post(url, headers=headers, json=body)
                return resp.status_code, resp.text

        status, text = await asyncio.to_thread(_post)
    except Exception as exc:
        return f"[failed] WhatsApp send request error: {exc}"

    if status >= 400:
        try:
            data = json.loads(text) if text else {}
            err = data.get("error") if isinstance(data, dict) else None
            if isinstance(err, dict):
                return f"[failed] {err.get('code', 'error')}: {err.get('message', text[:300])}"
        except json.JSONDecodeError:
            pass
        return f"[failed] HTTP {status}: {text[:300]}"

    return f"Sent {body['file_name']} to WhatsApp."


def openai_agents3_tool_definitions(
    identity: AgentIdentity,
    *,
    groups: tuple[str, ...],
    skill_filter: list[str] | None = None,
    instruction: str | None = None,
) -> list[dict[str, Any]]:
    """OpenAI tool schemas for hybrid local tools (gui-agents S3 stack)."""
    tools: list[dict[str, Any]] = []
    ids: list[str] = []
    if "files" in groups:
        ids.extend(_filter_tools(LOCAL_TOOL_IDS, identity, skill_filter, instruction=instruction))
    if "code" in groups and not is_read_only_file_intent(instruction or ""):
        for tid in _filter_tools(CODE_TOOL_IDS, identity, skill_filter, instruction=instruction):
            if tid not in ids:
                ids.append(tid)
    if "gui" in groups:
        for tid in _filter_tools(GUI_TOOL_IDS, identity, skill_filter, instruction=instruction):
            if tid not in ids:
                ids.append(tid)

    schemas: dict[str, dict[str, Any]] = {
        "s3_read_file": {
            "name": "s3_read_file",
            "description": (
                "Read a text file on the user's computer. Provide the full absolute path "
                "(e.g. C:\\Users\\...\\file.log on Windows). Use max_lines for large logs; "
                "prefer this over s3_python for read-only log review."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Absolute path to the file."},
                    "max_lines": {"type": "integer", "description": "Optional max lines to return."},
                },
                "required": ["path"],
            },
        },
        "s3_write_file": {
            "name": "s3_write_file",
            "description": "Write text to a file on the user's computer (overwrite).",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "content": {"type": "string"},
                },
                "required": ["path", "content"],
            },
        },
        "s3_list_dir": {
            "name": "s3_list_dir",
            "description": "List files and folders in a directory on the user's computer.",
            "parameters": {
                "type": "object",
                "properties": {"path": {"type": "string", "description": "Absolute directory path."}},
                "required": ["path"],
            },
        },
        "s3_open_file": {
            "name": "s3_open_file",
            "description": (
                "Open a file or folder on the user's computer with the default app or File Explorer "
                "(like remote desktop — windows appear on their screen, not in Qlix). "
                "Use mode 'folder' to open a directory in Explorer/Finder; 'reveal' to highlight a file in its folder."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Absolute path to file or folder."},
                    "mode": {
                        "type": "string",
                        "enum": ["default", "folder", "reveal"],
                        "description": "default=launch with associated app; folder=open directory; reveal=show in file manager.",
                    },
                    "application": {
                        "type": "string",
                        "description": "Optional app name or path (e.g. notepad, code, excel). Windows only.",
                    },
                },
                "required": ["path"],
            },
        },
        "s3_bash": {
            "name": "s3_bash",
            "description": (
                "Run a shell command via Agent-S3 LocalEnv (PowerShell on Windows, bash on Mac/Linux). "
                "Use for directory listings, pipelines, and system commands."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "command": {"type": "string"},
                    "timeout_seconds": {"type": "integer"},
                },
                "required": ["command"],
            },
        },
        "s3_python": {
            "name": "s3_python",
            "description": "Run a short Python snippet locally via Agent-S3 LocalEnv.",
            "parameters": {
                "type": "object",
                "properties": {"code": {"type": "string"}},
                "required": ["code"],
            },
        },
        "s3_create_pdf": {
            "name": "s3_create_pdf",
            "description": (
                "Create a real PDF file on the user's computer from text or markdown. "
                "Use this whenever the user asks to 'create/make/generate a PDF' or document — "
                "do NOT tell the user to copy-paste into Word. Supports markdown headings "
                "(#, ##, ###) and bullet lines ('- ' or '* '). Returns the absolute path of "
                "the created PDF, which you can then pass to s3_send_whatsapp_document."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "title": {"type": "string", "description": "Document title (rendered at the top)."},
                    "content": {
                        "type": "string",
                        "description": "Body text or markdown for the PDF.",
                    },
                    "output_path": {
                        "type": "string",
                        "description": (
                            "Optional absolute path for the .pdf. If omitted, it is saved to the "
                            "user's Documents folder using a slug of the title."
                        ),
                    },
                },
                "required": ["content"],
            },
        },
        "s3_create_xlsx": {
            "name": "s3_create_xlsx",
            "description": (
                "Create a real Excel .xlsx spreadsheet on the user's computer. Use this "
                "whenever the user asks for a spreadsheet/Excel/xlsx file — do NOT use "
                "s3_write_file for .xlsx (that produces a corrupt file Excel cannot open). "
                "Provide 'rows' as a list of lists (first row = headers) or a list of objects. "
                "Returns the absolute path, which you can pass to s3_send_whatsapp_document."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "title": {"type": "string", "description": "Used for the filename when output_path is omitted."},
                    "sheet_name": {"type": "string", "description": "Worksheet name (default 'Sheet1')."},
                    "rows": {
                        "type": "array",
                        "description": (
                            "Spreadsheet data: a list of rows where each row is a list of cell "
                            "values, OR a list of objects (keys become the header row)."
                        ),
                        "items": {},
                    },
                    "output_path": {
                        "type": "string",
                        "description": (
                            "Optional absolute path for the .xlsx. If omitted, saved to the user's "
                            "Documents folder using a slug of the title."
                        ),
                    },
                },
                "required": ["rows"],
            },
        },
        "s3_send_whatsapp_document": {
            "name": "s3_send_whatsapp_document",
            "description": (
                "Send a local file (e.g. a PDF created with s3_create_pdf) to the user's "
                "linked WhatsApp. Provide the absolute file_path. Use after creating a document "
                "when the user asked to send it on WhatsApp."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Absolute path to the file to send.",
                    },
                    "file_name": {
                        "type": "string",
                        "description": "Optional display name for the document in WhatsApp.",
                    },
                },
                "required": ["file_path"],
            },
        },
        "s3_code_task": {
            "name": "s3_code_task",
            "description": (
                "Run a multi-step coding subtask via Agent-S3 CodeAgent (bash/python). "
                "Use for spreadsheet/doc automation; requires local LLM API keys."
            ),
            "parameters": {
                "type": "object",
                "properties": {"task": {"type": "string"}},
                "required": ["task"],
            },
        },
        "gui_control": {
            "name": "gui_control",
            "description": (
                "Control desktop apps via Agent-S3 (screenshot + UI actions). "
                "Use for native apps; prefer s3_read_file / s3_bash for files."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "goal": {"type": "string"},
                    "max_steps": {"type": "integer"},
                },
                "required": ["goal"],
            },
        },
    }

    for tid in ids:
        spec = schemas.get(tid)
        if spec:
            tools.append({"type": "function", "function": spec})
    return tools


def build_agents3_executors(
    identity: AgentIdentity,
    *,
    groups: tuple[str, ...],
    qlix_sdk: Any,
    skill_filter: list[str] | None = None,
    agents3_context: Agents3RunContext | None = None,
    instruction: str | None = None,
) -> dict[str, Agents3Executor]:
    """Async executors; must run on the hybrid runner event loop."""
    executors: dict[str, Agents3Executor] = {}
    tool_ids: list[str] = []
    code_ids: tuple[str, ...] = () if is_read_only_file_intent(instruction or "") else CODE_TOOL_IDS
    if "files" in groups or code_ids or "gui" in groups:
        tool_ids = list(
            dict.fromkeys(
                _filter_tools(
                    tuple(
                        LOCAL_TOOL_IDS
                        + code_ids
                        + (GUI_TOOL_IDS if "gui" in groups else ())
                    ),
                    identity,
                    skill_filter,
                    instruction=instruction,
                )
            )
        )
    for tid in tool_ids:
        action_type = TOOL_SCOPE_MAP.get(tid, (tid,))[0]
        risk = "high" if tid in (
            "s3_bash",
            "s3_write_file",
            "s3_python",
            "s3_code_task",
            "gui_control",
        ) else "low"

        def _make(tid: str = tid, act: str = action_type, r: str = risk) -> Agents3Executor:
            async def _execute(args_json: str) -> str:
                try:
                    params = json.loads(args_json) if args_json.strip() else {}
                except json.JSONDecodeError:
                    params = {}
                if not isinstance(params, dict):
                    params = {}

                async def _run() -> str:
                    return await _dispatch_agents3_tool(
                        tid, params, identity=identity, agents3_context=agents3_context
                    )

                if qlix_sdk is None:
                    return await _run()
                run_id = (agents3_context.run_id if agents3_context else None) or ""
                payload: dict[str, Any] = {"tool": tid, **params}
                if run_id:
                    payload["runId"] = run_id
                result = await qlix_sdk.run(
                    act,
                    payload,
                    _run,
                    risk_level=r,
                )
                return str(result)

            return _execute

        executors[tid] = _make()

    return executors


async def _emit_agents3_log(
    agents3_context: Agents3RunContext | None,
    payload: dict[str, Any],
) -> None:
    if agents3_context is None or agents3_context.log_emit is None:
        return
    await agents3_context.log_emit(payload)


async def _dispatch_agents3_tool(
    tool_id: str,
    params: dict[str, Any],
    *,
    identity: AgentIdentity,
    agents3_context: Agents3RunContext | None = None,
) -> str:
    if tool_id in LOCAL_TOOL_IDS:
        await _emit_agents3_log(
            agents3_context,
            {
                "message": "agents3_step",
                "tool": tool_id,
                "phase": "execute",
                "detail": str(params.get("path") or params.get("command") or "")[:120],
            },
        )

    if tool_id == "s3_read_file":
        path = Path(str(params.get("path", "")))
        if not path.is_file():
            return f"[failed] File not found: {path}"
        text = path.read_text(encoding="utf-8", errors="replace")
        max_lines = params.get("max_lines")
        if max_lines and int(max_lines) > 0:
            text = "".join(text.splitlines(keepends=True)[: int(max_lines)])
        return text

    if tool_id == "s3_write_file":
        path = Path(str(params.get("path", "")))
        suffix = path.suffix.lower()
        if suffix in _BINARY_DOC_SUFFIXES:
            tool_hint = (
                "s3_create_pdf"
                if suffix == ".pdf"
                else "s3_create_xlsx"
                if suffix in (".xlsx", ".xls")
                else "a dedicated document tool"
            )
            return (
                f"[failed] Cannot write a {suffix} file with s3_write_file — it only writes "
                f"plain text, which produces a corrupt {suffix} the application cannot open. "
                f"Use {tool_hint} instead."
            )
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(str(params.get("content", "")), encoding="utf-8")
        return f"Wrote {path}"

    if tool_id == "s3_list_dir":
        path = Path(str(params.get("path", "")))
        if not path.is_dir():
            return f"[failed] Not a directory: {path}"
        entries = []
        for p in sorted(path.iterdir())[:200]:
            kind = "dir" if p.is_dir() else "file"
            entries.append(f"{kind}\t{p.name}")
        return "\n".join(entries) if entries else "(empty directory)"

    if tool_id == "s3_open_file":
        path = Path(str(params.get("path", "")))
        mode = str(params.get("mode") or "default")
        application = str(params.get("application") or "").strip() or None
        await _emit_agents3_log(
            agents3_context,
            {
                "message": "agents3_step",
                "tool": "s3_open_file",
                "phase": "open",
                "detail": f"{mode}: {path}",
            },
        )
        ok, msg = _open_path_on_system(path, mode=mode, application=application)
        return msg if ok else f"[failed] {msg}"

    if tool_id == "s3_bash":
        timeout = int(params.get("timeout_seconds") or 120)
        result = _run_bash(str(params.get("command", "")), timeout=timeout)
        out = result.get("output") or result.get("error") or ""
        prefix = "" if result.get("status") == "ok" else "[failed] "
        return prefix + str(out)

    if tool_id == "s3_python":
        result = _run_python(str(params.get("code", "")))
        out = (result.get("output") or "") + (result.get("error") or "")
        prefix = "" if result.get("status") == "ok" else "[failed] "
        return prefix + str(out)

    if tool_id == "s3_create_pdf":
        ok, msg = _create_pdf_file(
            title=str(params.get("title") or ""),
            content=str(params.get("content") or ""),
            output_path=str(params.get("output_path") or "").strip() or None,
        )
        await _emit_agents3_log(
            agents3_context,
            {
                "message": "agents3_step",
                "tool": "s3_create_pdf",
                "phase": "write",
                "detail": msg[:160],
            },
        )
        return msg if ok else f"[failed] {msg}"

    if tool_id == "s3_create_xlsx":
        ok, msg = _create_xlsx_file(
            title=str(params.get("title") or ""),
            rows=params.get("rows"),
            sheet_name=str(params.get("sheet_name") or "Sheet1"),
            output_path=str(params.get("output_path") or "").strip() or None,
        )
        await _emit_agents3_log(
            agents3_context,
            {
                "message": "agents3_step",
                "tool": "s3_create_xlsx",
                "phase": "write",
                "detail": msg[:160],
            },
        )
        return msg if ok else f"[failed] {msg}"

    if tool_id == "s3_send_whatsapp_document":
        file_path = str(params.get("file_path") or "").strip()
        file_name = str(params.get("file_name") or "").strip() or None
        await _emit_agents3_log(
            agents3_context,
            {
                "message": "agents3_step",
                "tool": "s3_send_whatsapp_document",
                "phase": "send",
                "detail": file_path[:160],
            },
        )
        return await _send_whatsapp_document(
            file_path,
            file_name,
            identity=identity,
            agents3_context=agents3_context,
        )

    if tool_id == "s3_code_task":
        await _emit_agents3_log(
            agents3_context,
            {
                "message": "agents3_step",
                "tool": "s3_code_task",
                "phase": "start",
                "detail": (str(params.get("task", ""))[:120] or "CodeAgent task"),
            },
        )
        return await _run_s3_code_task(
            str(params.get("task", "")),
            identity=identity,
            agents3_context=agents3_context,
        )

    if tool_id == "gui_control":
        await _emit_agents3_log(
            agents3_context,
            {
                "message": "agents3_step",
                "tool": "gui_control",
                "phase": "start",
                "detail": (str(params.get("goal", ""))[:120] or "Desktop control"),
            },
        )
        return await _run_gui_control(
            str(params.get("goal", "")),
            int(params.get("max_steps") or 8),
            identity=identity,
            agents3_context=agents3_context,
        )

    return f"[failed] Unknown tool: {tool_id}"


def _s3_run_context(
    identity: AgentIdentity,
    agents3_context: Agents3RunContext | None,
) -> tuple[str, str, str, str]:
    agent_id = agents3_context.agent_id if agents3_context else identity.agent_id
    runner_token = agents3_context.runner_token if agents3_context else ""
    model = (
        agents3_context.model
        if agents3_context and agents3_context.model.strip()
        else os.environ.get("QLIX_PROXY_MODEL", "openrouter/openai/gpt-4o-mini")
    )
    return identity.backend_url, agent_id, runner_token, model


async def _run_s3_code_task(
    task: str,
    *,
    identity: AgentIdentity,
    agents3_context: Agents3RunContext | None,
) -> str:
    if not task.strip():
        return "[failed] task is required"
    try:
        import os

        from gui_agents.s3.agents.code_agent import CodeAgent
        from gui_agents.s3.utils.local_env import LocalEnv
    except ImportError:
        return "[failed] pip install gui-agents (Agent-S3)"

    backend_url, agent_id, runner_token, model = _s3_run_context(identity, agents3_context)
    engine_params, _ = resolve_s3_engine_params(
        identity.llm_mode,
        backend_url=backend_url,
        agent_id=agent_id,
        runner_token=runner_token,
        model=model,
    )
    env = LocalEnv()
    agent = CodeAgent(engine_params=engine_params, budget=int(os.environ.get("QLIX_S3_CODE_BUDGET", "8")))
    try:
        result = await asyncio.to_thread(agent.execute, task, "", env.controller)
    except Exception as exc:
        msg = str(exc)
        if "inference_not_configured" in msg or "OPENROUTER" in msg.upper():
            return (
                "[failed] Qlix inference proxy is not configured on the server "
                "(OPENROUTER_API_KEY). Contact your administrator."
            )
        return f"[failed] {msg}"
    summary = result.get("summary") or result.get("completion_reason") or str(result)
    return str(summary)


async def _run_gui_control(
    goal: str,
    max_steps: int,
    *,
    identity: AgentIdentity,
    agents3_context: Agents3RunContext | None,
) -> str:
    if not goal.strip():
        return "[failed] goal is required"
    try:
        import os

        import pyautogui  # type: ignore[import-untyped]
        from gui_agents.s3.agents.agent_s import AgentS3
        from gui_agents.s3.agents.grounding import OSWorldACI
    except ImportError:
        return "[failed] pip install gui-agents pyautogui for desktop control"

    backend_url, agent_id, runner_token, model = _s3_run_context(identity, agents3_context)
    grounding_model = os.environ.get("QLIX_S3_GROUNDING_MODEL", "").strip() or None
    engine_params, grounding_params = resolve_s3_engine_params(
        identity.llm_mode,
        backend_url=backend_url,
        agent_id=agent_id,
        runner_token=runner_token,
        model=model,
        grounding_model=grounding_model,
    )
    grounding = OSWorldACI(
        env=None,
        platform=platform.system().lower(),
        engine_params_for_generation=engine_params,
        engine_params_for_grounding=grounding_params,
        width=int(os.environ.get("QLIX_S3_SCREEN_WIDTH", "1920")),
        height=int(os.environ.get("QLIX_S3_SCREEN_HEIGHT", "1080")),
    )
    agent = AgentS3(engine_params, grounding, platform=platform.system().lower())
    max_steps = max(1, min(max_steps, 20))
    last_actions: list[str] = []

    try:
        for step_idx in range(max_steps):
            await _emit_agents3_log(
                agents3_context,
                {
                    "message": "agents3_step",
                    "tool": "gui_control",
                    "phase": "predict",
                    "step": step_idx + 1,
                    "max_steps": max_steps,
                    "detail": "Screenshot + UI action",
                },
            )
            screenshot = pyautogui.screenshot()
            buf = __import__("io").BytesIO()
            screenshot.save(buf, format="PNG")
            import base64

            obs = {"screenshot": base64.b64encode(buf.getvalue()).decode("ascii")}
            _, actions = await asyncio.to_thread(agent.predict, goal, obs)
            last_actions = actions or []
            if any("done" in str(a).lower() or "fail" in str(a).lower() for a in last_actions):
                break
    except Exception as exc:
        msg = str(exc)
        if "inference_not_configured" in msg or "OPENROUTER" in msg.upper():
            return (
                "[failed] Qlix inference proxy is not configured on the server "
                "(OPENROUTER_API_KEY). Contact your administrator."
            )
        return f"[failed] {msg}"

    return "Agent-S3 steps completed. Last actions: " + "; ".join(map(str, last_actions[:5]))
