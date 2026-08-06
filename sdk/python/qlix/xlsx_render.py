"""Shared spreadsheet (.xlsx) renderer (openpyxl).

Used by the hybrid file tool (``agents3_runtime.luna_local_create_xlsx``) and the cloud
document tool (``cloud_document_runtime.create_xlsx``) so both produce identical
workbooks.
"""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any


def coerce_rows(data: Any) -> list[list[Any]] | None:
    """Normalize accepted xlsx inputs into a list of rows (list of cells)."""
    if isinstance(data, str):
        text = data.strip()
        if not text:
            return None
        try:
            parsed = json.loads(text)
        except (json.JSONDecodeError, ValueError):
            parsed = None
        if parsed is not None:
            return coerce_rows(parsed)
        return [row for row in csv.reader(io.StringIO(text))]
    if isinstance(data, list):
        if not data:
            return None
        if all(isinstance(r, dict) for r in data):
            headers: list[str] = []
            for r in data:
                for k in r.keys():
                    if k not in headers:
                        headers.append(str(k))
            rows: list[list[Any]] = [headers]
            for r in data:
                rows.append([r.get(h, "") for h in headers])
            return rows
        out: list[list[Any]] = []
        for r in data:
            if isinstance(r, list):
                out.append(list(r))
            else:
                out.append([r])
        return out
    return None


def render_xlsx(
    *,
    title: str,
    rows: Any,
    sheet_name: str,
    out: Path,
) -> tuple[bool, str]:
    """Write a real .xlsx workbook to ``out``.

    ``rows`` may be a list of lists, a list of dicts (keys become the header row),
    a JSON string of either, or CSV text. Returns ``(ok, message)``.
    """
    table = coerce_rows(rows)
    if not table:
        return False, "rows is required to create a spreadsheet (list of lists, list of dicts, JSON, or CSV)"

    if out.suffix.lower() != ".xlsx":
        out = out.with_suffix(".xlsx")

    try:
        out.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        return False, f"Could not create output directory: {exc}"

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return (
            False,
            "openpyxl is not installed. Install the hybrid extras: "
            "pip install 'qlix[hybrid]' (or pip install openpyxl).",
        )

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = (sheet_name or "Sheet1")[:31]
        for r_idx, row in enumerate(table, start=1):
            for c_idx, cell in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=cell)
        if table:
            for c_idx in range(1, len(table[0]) + 1):
                ws.cell(row=1, column=c_idx).font = Font(bold=True)
            for c_idx in range(1, max(len(r) for r in table) + 1):
                width = max(
                    (len(str(r[c_idx - 1])) for r in table if c_idx - 1 < len(r)),
                    default=10,
                )
                ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = min(
                    max(width + 2, 10), 60
                )
        wb.save(str(out))
    except Exception as exc:
        return False, f"XLSX generation error: {exc}"

    label = title.strip() or out.stem
    return True, f"Created spreadsheet: {out.resolve()} ({label})"
